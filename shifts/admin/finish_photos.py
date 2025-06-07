import io

import httpx
import openpyxl
from django.contrib import admin
from django.db.models import QuerySet
from django.http.request import HttpRequest
from django.http.response import HttpResponse
from django.utils.translation import gettext_lazy as _
from import_export import resources
from import_export.admin import ImportExportModelAdmin
from openpyxl.drawing.image import Image as OpenPyxlImage, PILImage
from openpyxl.utils.cell import get_column_letter

from shifts.mixins import ShiftModelStaffSelectRelatedMixin
from shifts.models import Shift, ShiftFinishPhoto


class ShiftFinishPhotoResource(resources.ModelResource):
    class Meta:
        model = ShiftFinishPhoto


class HasUrlFilter(admin.SimpleListFilter):
    title = _("Has url")
    parameter_name = "has_url"

    def lookups(self, request, model_admin):
        return (
            ("true", _("yes")),
            ("false", _("no")),
        )

    def queryset(self, request, queryset: QuerySet):
        if self.value() == "true":
            return queryset.exclude(url__isnull=True)
        if self.value() == "false":
            return queryset.filter(url__isnull=True)
        return queryset


@admin.action(description=_("Export excel file with photos"))
def download_xlsx(
        modeladmin: 'ShiftFinishPhotoAdmin',
        request: HttpRequest,
        queryset: QuerySet[ShiftFinishPhoto],
):
    queryset = queryset.select_related(
        "shift",
        "shift__staff",
        "shift__car_wash",
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MyModel Data"

    headers = [
        'ID',
        'URL фотографии',
        'Дата смены',
        'ФИО сотрудника',
        'Адрес мойки',
        'Фотография',
    ]
    ws.append(headers)

    row_num = 2

    with httpx.Client() as http_client:
        for obj in queryset:
            obj: ShiftFinishPhoto

            # Write ID and URL
            ws.cell(row=row_num, column=1, value=obj.id)
            ws.cell(row=row_num, column=2, value=obj.url)
            ws.cell(row=row_num, column=3, value=obj.shift.date)
            ws.cell(row=row_num, column=4, value=obj.shift.staff.full_name)
            ws.cell(row=row_num, column=5, value=obj.shift.car_wash.name)

            try:
                img_response = http_client.get(obj.url)
                img_response.raise_for_status()

                # Validate and load image
                img_pil = PILImage.open(io.BytesIO(img_response.content))
                img_width_px, img_height_px = img_pil.size

                # Save image as PNG in memory
                img_io = io.BytesIO()
                img_pil.save(img_io, format='JPEG')
                img_io.seek(0)

                # Insert image
                excel_img = OpenPyxlImage(img_io)
                ws.add_image(excel_img, f"C{row_num}")

                # --- Resize row and column to fit image --- #

                # Convert pixels to Excel units
                # Excel row height unit = ~0.75 points per pixel
                # Excel column width unit ≈ pixel_width / 7 (approx.)

                row_height_points = img_height_px * 0.75
                col_width = img_width_px / 7.0

                ws.row_dimensions[row_num].height = row_height_points

                col_letter = get_column_letter(6)
                ws.column_dimensions[col_letter].width = col_width

            except Exception:
                ws.cell(row=row_num, column=6, value="Image Error")

            row_num += 1

    # Optional: widen ID and URL columns
    ws.column_dimensions[get_column_letter(1)].width = 15
    ws.column_dimensions[get_column_letter(2)].width = 40
    ws.column_dimensions[get_column_letter(3)].width = 15
    ws.column_dimensions[get_column_letter(4)].width = 25
    ws.column_dimensions[get_column_letter(5)].width = 30

    # Save to stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return HttpResponse(
        stream.read(),
        content_type='application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; '
                                   'filename="shift_finish_photos.xlsx"'
        }
    )


@admin.register(ShiftFinishPhoto)
class ShiftFinishPhotoAdmin(
    ShiftModelStaffSelectRelatedMixin,
    ImportExportModelAdmin,
):
    resource_class = ShiftFinishPhotoResource
    list_display = ("shift", "url")
    list_select_related = ("shift", "shift__staff")
    list_filter = ("shift__car_wash", HasUrlFilter)
    actions = [download_xlsx]
    autocomplete_fields = ("shift",)
